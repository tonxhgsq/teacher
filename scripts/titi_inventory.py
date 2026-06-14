#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ROOT = Path("/Users/tang/Desktop/titi")
OUT = Path("tmp/titi_inventory.json")

QUESTION_EXTS = {".pdf", ".csv", ".xlsx"}
NON_QUESTION_KEYWORDS = [
    "答案",
    "解析",
    "详解",
    "点拨",
    "板书",
]
SUPPLEMENT_KEYWORDS = ["公式", "知识点"]


def strip_ext(name: str) -> str:
    return re.sub(r"\.[^.]+$", "", name)


def normalize_title(path: Path) -> str:
    title = strip_ext(path.name)
    title = re.sub(r"\s+", " ", title).strip()
    title = title.replace("【A4】", "").replace("【A4答案】", "")
    title = title.replace("A4-", "").replace("A4", "")
    title = re.sub(r"[-_ ]*题目$", "", title)
    title = re.sub(r"[-_ ]*(答案|解析|详解|与名校点拨|名校点拨|板书)$", "", title)
    title = title.replace("（解析）", "").replace("【解析】", "")
    title = re.sub(r"\s+", " ", title).strip()
    return title


def classify_file(path: Path) -> str:
    name = path.name
    if any(k in name for k in NON_QUESTION_KEYWORDS):
        return "answer_or_solution"
    if any(k in name for k in SUPPLEMENT_KEYWORDS):
        return "supplement_maybe"
    return "question"


def read_csv_header(path: Path) -> list[str]:
    for enc in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            with path.open("r", encoding=enc, newline="") as f:
                reader = csv.reader(f)
                return next(reader, [])
        except Exception:
            continue
    return []


def read_xlsx_header(path: Path) -> list[str]:
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        return [str(v).strip() if v is not None else "" for v in row]
    except Exception:
        return []


def count_structured_rows(path: Path) -> int:
    try:
        if path.suffix.lower() == ".csv":
            for enc in ("utf-8-sig", "utf-8", "gb18030"):
                try:
                    with path.open("r", encoding=enc, newline="") as f:
                        rows = list(csv.reader(f))
                    return max(0, len(rows) - 1)
                except Exception:
                    continue
        if path.suffix.lower() == ".xlsx":
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            return max(0, ws.max_row - 1)
    except Exception:
        return 0
    return 0


def main() -> None:
    files = [p for p in ROOT.rglob("*") if p.is_file() and p.suffix.lower() in QUESTION_EXTS]
    records = []
    groups: dict[str, list[dict]] = defaultdict(list)
    for path in sorted(files):
        kind = classify_file(path)
        title = normalize_title(path)
        rel = str(path.relative_to(ROOT))
        row = {
            "path": str(path),
            "relative": rel,
            "name": path.name,
            "ext": path.suffix.lower().lstrip("."),
            "kind": kind,
            "title": title,
            "rows": count_structured_rows(path) if path.suffix.lower() in {".csv", ".xlsx"} else None,
        }
        if path.suffix.lower() == ".csv":
            row["header"] = read_csv_header(path)
        elif path.suffix.lower() == ".xlsx":
            row["header"] = read_xlsx_header(path)
        records.append(row)
        groups[f"{path.parent}:{title}"].append(row)

    question_groups = []
    for key, items in groups.items():
        question_items = [x for x in items if x["kind"] == "question"]
        if not question_items:
            continue
        by_ext = defaultdict(list)
        for item in items:
            by_ext[item["ext"]].append(item)
        title = question_items[0]["title"]
        # Prefer structured question data when present, but keep PDF for verification.
        preferred = None
        for ext in ("csv", "xlsx", "pdf"):
            candidates = [x for x in question_items if x["ext"] == ext]
            if candidates:
                preferred = candidates[0]
                break
        question_groups.append({
            "title": title,
            "directory": str(Path(question_items[0]["path"]).parent),
            "preferred": preferred,
            "question_files": question_items,
            "all_files": items,
            "has_pdf": bool(by_ext["pdf"]),
            "has_csv": bool(by_ext["csv"]),
            "has_xlsx": bool(by_ext["xlsx"]),
        })

    summary = {
        "root": str(ROOT),
        "total_files": len(records),
        "by_ext": Counter(r["ext"] for r in records),
        "by_kind": Counter(r["kind"] for r in records),
        "question_groups": len(question_groups),
        "groups_with_pdf": sum(1 for g in question_groups if g["has_pdf"]),
        "groups_with_structured": sum(1 for g in question_groups if g["has_csv"] or g["has_xlsx"]),
    }
    payload = {
        "summary": {
            **summary,
            "by_ext": dict(summary["by_ext"]),
            "by_kind": dict(summary["by_kind"]),
        },
        "records": records,
        "question_groups": question_groups,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(payload["summary"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
